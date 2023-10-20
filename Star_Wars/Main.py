from start import *
#from music import *
import pygame
import sys
import openpyxl

def add_values_to_excel(file_path, sheet_name, values):
    try:
        try:
            workbook = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)

        sheet = workbook[sheet_name]

        for row in values:
            sheet.append(row)

        workbook.save(file_path)

    except Exception as e:
        print(f"An error occurred: {e}")

def help():
    print('''\nBASIC:\nYour opponent will not be automatic\n\nChoose your character as P1\n
        sometimes a character may undergo aggeression and may\ncause mor damage than normal\n
        more featres will be added in future\n
        ''')
    
def run(name):
    play_music()
    print('type p and press enter to play the game\nh for help\ne for exit\n')
    choise = input()
    if choise == 'p':
        Game(name)
    elif choise == 'h':
        help()
    elif choise == 'e':
        sys.exit()
    else:
        print('Input Error : choise made was invalid\n')

def mainF():
    if __name__ == '__main__':
        user = input('type s to sign up\nl to login\n')
        if user == 'login' or user == 'l':
            name = input('Input your name jedi: ')
            password = input('Input password: ')
            if name == password:
                run(name)
            else:
                print('username or password is incorrect')
        elif user == 'sign up' or user == 's':
            name = input('Please enter your name: ')
            password = input('Please enter your password: ')
            password2 = input('Please enter your confirm fassword: ')
            age = input('Please enter your age: ')
            gender = input('Please enter your gender: ')
            excel_file_path = "data.xlsx"
            sheet_name = "DataSheet"
            new_values = [
                ["Name","Password",'age','gender'],
                [name,password,age,gender]
            ]
            add_values_to_excel(excel_file_path, sheet_name, new_values)
            if len(password) > 7 and password == password2 and gender == 'male' or gender == 'female' or gender == 'others':
                run(name)

mainF()
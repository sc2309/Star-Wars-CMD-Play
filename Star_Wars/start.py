import random
from Maps import *
from strike import *
from Characters import *
#from music import *
import openpyxl

def add_values_to_excel(file_path, sheet_name, values):
    try:
        try:
            workbook = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)

        sheet = workbook[sheet_name]

        for row in values:
            sheet.append(row)

        workbook.save(file_path)

    except Exception as e:
        print(f"An error occurred: {e}")

def Game(name):
        play_music()
        print('Welcome to Star Jedi Fighters!')
        Galactic_Credit = 0
        map = str(input('Please select map from here:\nCoruscant\nMustafar\nStar Destroyer\nStar Ship\nNaboo\ntatooine\nKamino\nHoth\nExegol\nMandalore\n'))
        Player1 = input('\n\nHere you can select a character for yourself\n\nKylo Ren 500\nDarth Malgus 550\nPlapertine 700\nRay Skywalker 500\nLuke Skywalker 650\nObi Van Kenobi 650\nYoda 1000\nAnakin Skywalker or Darth Vader 700\nQui Gon Jinn 650\nDarth Maul 620\nDarth Revan 650\nCount Dooku 630\n')
        player2 = input('Let your opponent select his character as P2:\n\nKylo Ren 500\nDarth Malgus 550\nPlapertine 700\nRay Skywalker 500\nLuke Skywalker 650\nObi Van Kenobi 650\nYoda 1000\nAnakin Skywalker or Darth Vader 700\nQui Gon Jinn 650\nDarth Maul 620\nDarth Revan 650\nCount Dooku 630\n')
        if Player1 == 'Kylo Ren':
            h,sh,shand,sb,sl,c = Kylo_Ren()
        elif Player1 == 'Darth Malgus':
            h,sh,shand,sb,sl,c = Malgus()
        elif Player1 == 'Plapertine':
            h,sh,shand,sb,sl,c = Plapertine()
        elif Player1 == 'Ray Skywalker':
            h,sh,shand,sb,sl,c = Ray()
        elif Player1 == 'Luke Skywalker':
            h,sh,shand,sb,sl,c = luke()
        elif Player1 == 'Obi Van Kenobi':
            h,sh,shand,sb,sl,c = ObiVan()
        elif Player1 == 'Yoda':
            h,sh,shand,sb,sl,c = Yoda()
        elif Player1 == 'Anakin Skywalker' or Player1 == 'Darth Vader':
            h,sh,shand,sb,sl,c = anakin()
        elif Player1 == 'Qui Gon Jinn':
            h,sh,shand,sb,sl,c = QuiGonJinn()
        elif Player1 == 'Darth Maul':
            h,sh,shand,sb,sl,c = maul()
        elif Player1 == 'Darth Revan':
            h,sh,shand,sb,sl,c = DarthRevan()
        elif Player1 == 'Count Dooku':
            h,sh,shand,sb,sl,c = countDooku()
        else:
            print('ERROR 1 : Not a character')

        if player2 == 'Kylo Ren':
            h2,sh2,shand2,sb2,sl2,c2 = Kylo_Ren()
        elif player2 == 'Darth Malgus':
            h2,sh2,shand2,sb2,sl2,c2 = Malgus()
        elif player2 == 'Plapertine':
            h2,sh2,shand2,sb2,sl2,c2 = Plapertine()
        elif player2 == 'Ray Skywalker':
            h2,sh2,shand2,sb2,sl2,c2 = Ray()
        elif player2 == 'Luke Skywalker':
            h2,sh2,shand2,sb2,sl2,c2 = luke()
        elif player2 == 'Obi Van Kenobi':
            h2,sh2,shand2,sb2,sl2,c2 = ObiVan()
        elif player2 == 'Yoda':
            h2,sh2,shand2,sb2,sl2,c2 = Yoda()
        elif player2 == 'Anakin Skywalker' or player2 == 'Darth Vader':
            h2,sh2,shand2,sb2,sl2,c2 = anakin()
        elif player2 == 'Qui Gon Jinn':
            h2,sh2,shand2,sb2,sl2,c2 = QuiGonJinn()
        elif player2 == 'Darth Maul':
            h2,sh2,shand2,sb2,sl2,c2 = maul()
        elif player2 == 'Darth Revan':
            h2,sh2,shand2,sb2,sl2,c2 = DarthRevan()
        elif player2 == 'Count Dooku':
            h2,sh2,shand2,sb2,sl2,c2 = countDooku()
        else:
            print('ERROR 1 : Not a character')

        if map == 'Coruscant':
            pass
        elif map == 'Mustafar':
            pass
        elif map == 'Star Destroyer':
            pass
        elif map == 'Star Ship':
            pass
        elif map == 'Naboo':
            pass
        elif map == 'Tatooine':
            pass
        elif map == 'Kamino':
            pass
        elif map == 'Hoth':
            pass
        elif map == 'Exegol':
            pass
        elif map == 'Mandalore':
            pass
        else:
            pass

        excel_file_path = "Money.xlsx"
        sheet_name = "DataSheet"
        new_values = [
            ["Name","Money"],
            [name,Galactic_Credit]
        ]
        add_values_to_excel(excel_file_path, sheet_name, new_values)

        x = random.randrange(1, 3)
        if x == 1:
            print("Player 1 got first chance\n\n")
            P1Strike(h,sh,shand,sb,sl,c,h2,sh2,shand2,sb2,sl2,c2)
        else:
            print("Player 2 got first chance\n\n")
            P2Strike(h,sh,shand,sb,sl,c,h2,sh2,shand2,sb2,sl2,c2)